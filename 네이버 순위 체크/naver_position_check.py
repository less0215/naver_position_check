import streamlit as st
import time
import random
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Streamlit 앱 제목
st.title("네이버 순위 체크")

# 팀 선택
selected_team = st.selectbox("팀 선택", ["청소년팀", "형사팀"])

# 키워드 입력
keywords = st.text_area("키워드를 입력해 주세요 (한 줄에 하나씩)", height=200)

# 청소년팀 ID 리스트
dongju_id_list = [
    "designersiun", "singsong0514", "phoenixjeong", "hamas3000", "roses777",
    "dongjulaw1", "dongjulaw2", "dongjusuwon1", "dongjulaw6", "dj_ehdwn1",
    "rudnfdldi00", "ehtlarhdwn", "widance", "yellowoi", "dongjulaw",
    "tale1396", "dongjulaw5", "dongjulaw100", "dongjulaw4", "dongjulaw02",
    "dksro018", "cckjjt", "qusghtkehdwn", "dongjulaw7", "ujm159",
    "dong-ju-law", "dongjulaw3", "ehdwnfh", "kkobugi39"
]

# 결과를 실시간으로 표시하기 위한 함수
def color_keyword(val, keyword_types, keyword):
    keyword_type = keyword_types.get(keyword, '')
    if keyword_type == 'knowledge_snippet':
        return 'background-color: #90EE90'  # 밝은 초록색
    elif keyword_type == 'smartblock':
        return 'background-color: #ADD8E6'  # 밝은 파란색
    return ''

# 엑셀 파일 생성 함수
def create_excel(df, keyword_types):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # 헤더 추가
    for col, value in enumerate(df.columns.values, start=1):
        sheet.cell(row=1, column=col, value=value)

    # 데이터 추가 및 스타일 적용
    for row, (index, data) in enumerate(df.iterrows(), start=2):
        keyword = data['키워드']
        keyword_type = keyword_types.get(keyword, '')
        for col, value in enumerate(data.values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if col == 1:  # 키워드 열
                if keyword_type == 'knowledge_snippet':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif keyword_type == 'smartblock':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # 열 너비 자동 조정
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output)
    return output.getvalue()

# 검색 시작 버튼
if st.button("검색 시작"):
    if not keywords:
        st.error("키워드를 입력해주세요.")
    else:
        # 키워드 리스트 생성
        keyword_list = [keyword.strip() for keyword in keywords.split('\n') if keyword.strip()]
        
        if not keyword_list:
            st.error("유효한 키워드를 입력해주세요.")
        else:
            # Chrome 옵션 설정
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # 헤드리스 모드

            # WebDriver 초기화
            driver = webdriver.Chrome(options=chrome_options)

            # 결과를 저장할 리스트 초기화
            results_list = []
            keyword_types = {}  # 키워드 유형을 저장할 딕셔너리
            
            # 실시간 결과 표시를 위한 placeholder
            result_placeholder = st.empty()

            # 진행 상황 표시를 위한 progress bar
            progress_bar = st.progress(0)

            # 각 키워드에 대해 검색 수행
            for i, keyword in enumerate(keyword_list):
                # 검색 페이지로 이동
                driver.get(f"https://search.naver.com/search.naver?ssc=tab.nx.all&where=nexearch&sm=tab_jum&query={keyword}")

                try:
                    keyword_type = ''

                    # 지식스니펫 확인
                    try:
                        knowledge_snippet = driver.find_element(By.CSS_SELECTOR, '.source_box .txt.elss')
                        keyword_type = 'knowledge_snippet'
                    except:
                        pass

                    # 스마트블럭 확인
                    if not keyword_type:
                        soup = BeautifulSoup(driver.page_source, 'html.parser')
                        try:
                            smartblock_research = soup.select_one('.BZppu7wV32H2scXPRUVx.fds-info-inner-text')
                            keyword_type = 'smartblock'
                        except:
                            pass

                    # 키워드 유형 저장
                    keyword_types[keyword] = keyword_type

                    # 블로그 탭 클릭
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, '.flick_bx:nth-of-type(3) > a'))
                    ).click()
                    
                    # 무한스크롤 처리
                    last_height = driver.execute_script("return document.body.scrollHeight")
                    while True:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        time.sleep(random.uniform(1, 1.5))
                        new_height = driver.execute_script("return document.body.scrollHeight")
                        if new_height == last_height:
                            break
                        last_height = new_height

                    # 블로그 순위 체크
                    blog_ids = driver.find_elements(By.CSS_SELECTOR, '.user_info a')
                    results = {j: '' for j in range(1, 16)}  # 모든 순위를 빈 문자열로 초기화
                    for blog_id in blog_ids:
                        href = blog_id.get_attribute('href')
                        extracted_id = href.split('/')[-1]
                        if extracted_id in dongju_id_list:
                            index = dongju_id_list.index(extracted_id) + 1
                            if index <= 15:  # 15위까지만 저장
                                results[index] = extracted_id

                    # 결과 리스트에 추가
                    row = {'키워드': keyword, 'M': '', 'P': ''}
                    row.update(results)
                    results_list.append(row)

                    # 실시간으로 결과 표시
                    df = pd.DataFrame(results_list)
                    
                    # 키워드 열에만 배경색 적용
                    styled_df = df.style.apply(lambda row: [color_keyword(val, keyword_types, row['키워드']) if idx == 0 else '' for idx, val in enumerate(row)], axis=1)
                    
                    result_placeholder.dataframe(styled_df, width=1000)  # 너비 조정

                    # 진행 상황 업데이트
                    progress_bar.progress((i + 1) / len(keyword_list))

                except Exception as e:
                    st.error(f"키워드 '{keyword}' 검색 중 오류 발생: {str(e)}")

            driver.quit()

            # Progress bar 제거
            progress_bar.empty()

            # 엑셀 다운로드 버튼
            excel_data = create_excel(df, keyword_types)
            st.download_button(
                label="엑셀 다운로드",
                data=excel_data,
                file_name="search_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

st.info("'검색 시작' 버튼을 클릭하여 검색 결과를 분석하세요.")