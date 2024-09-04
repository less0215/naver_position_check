import streamlit as st
import time
import random
import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import hashlib
import hmac
import base64

st.set_page_config(
    page_title="법무법인 동주 SEO",
    layout='wide'
)

# Naver API 관련 함수 및 설정
BASE_URL = 'https://api.naver.com'
API_KEY = 'API_KEY'
SECRET_KEY = 'SECRET_KEY'
CUSTOMER_ID = 'CUSTOMER_ID'

class Signature:
    @staticmethod
    def generate(timestamp, method, uri, secret_key):
        message = "{}.{}.{}".format(timestamp, method, uri)
        hash = hmac.new(bytes(secret_key, "utf-8"), bytes(message, "utf-8"), hashlib.sha256)
        
        hash.hexdigest()
        return base64.b64encode(hash.digest())

def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(round(time.time() * 1000))
    signature = Signature.generate(timestamp, method, uri, secret_key)
    
    return {
        'Content-Type': 'application/json; charset=UTF-8',
        'X-Timestamp': timestamp,
        'X-API-KEY': api_key,
        'X-Customer': str(customer_id),
        'X-Signature': signature
    }

def get_search_volume(keyword):
    uri = '/keywordstool'
    method = 'GET'
    params = {'hintKeywords': keyword, 'showDetail': '1'}
    
    r = requests.get(BASE_URL + uri, params=params, 
                     headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID))
    
    data = r.json()['keywordList']
    result = next((item for item in data if item['relKeyword'] == keyword), None)
    
    if result:
        return result['monthlyPcQcCnt'], result['monthlyMobileQcCnt']
    else:
        return 0, 0

# 색상 적용 함수
def color_keyword(val, keyword_types, keyword):
    keyword_type = keyword_types.get(keyword, '')
    if keyword_type == 'knowledge_snippet':
        return 'background-color: #90EE90'  # 밝은 초록색
    elif keyword_type == 'smartblock':
        return 'background-color: #ADD8E6'  # 밝은 파란색
    elif keyword_type == 'both':
        return 'background-color: #FFB3BA'  # 밝은 빨간색
    return ''

# 엑셀 파일 생성 함수
def create_excel(df, keyword_types, smartblock_keywords):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    # 헤더 추가
    for col, value in enumerate(df.columns.values, start=1):
        sheet.cell(row=1, column=col, value=value)

    # 데이터 추가 및 스타일 적용
    for row, (index, data) in enumerate(df.iterrows(), start=2):
        keyword = data['키워드']
        keyword_type = keyword_types.get(keyword, '')
        for col, value in enumerate(data.values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            if col == 1:  # 키워드 열
                if keyword_type == 'knowledge_snippet':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif keyword_type == 'smartblock':
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif keyword_type == 'both':
                    cell.fill = PatternFill(start_color="FFB3BA", end_color="FFB3BA", fill_type="solid")

    # 스마트블럭 키워드 및 연관 키워드 추가
    sheet = workbook.create_sheet(title="스마트블럭 키워드")
    sheet.append(["스마트블럭 키워드", "연관 키워드"])
    for keyword, related_keywords in smartblock_keywords.items():
        sheet.append([keyword, ", ".join(related_keywords)])

    # 열 너비 자동 조정
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(output)
    return output.getvalue()

# 사이드탭 생성
selected_tab = st.sidebar.radio("검색 엔진 선택", ["네이버", "구글"])

if selected_tab == "네이버":
    # 네이버 탭 내용
    st.title("네이버 순위 체크 및 검색량 조회")

    # 팀 선택
    selected_team = st.selectbox("팀 선택", ["청소년팀", "형사팀"])

    # 키워드 입력
    keywords = st.text_area("키워드를 입력해 주세요 (한 줄에 하나씩)", height=200)

    # 청소년팀 ID 리스트
    dongju_id_list = [
        "designersiun", "singsong0514", "phoenixjeong", "hamas3000", "roses777",
        "dongjulaw1", "dongjulaw2", "dongjusuwon1", "dongjulaw6", "dj_ehdwn1",
        "rudnfdldi00", "ehtlarhdwn", "widance", "yellowoi", "dongjulaw",
        "tale1396", "dongjulaw5", "dongjulaw100", "dongjulaw4", "dongjulaw02",
        "dksro018", "cckjjt", "qusghtkehdwn", "dongjulaw7", "ujm159",
        "dong-ju-law", "dongjulaw3", "ehdwnfh", "kkobugi39"
    ]


    # 순위 확인 버튼
    if st.button("순위 확인"):
        if not keywords:
            st.error("키워드를 입력해주세요.")
        else:
            # 키워드 리스트 생성
            keyword_list = [keyword.strip() for keyword in keywords.split('\n') if keyword.strip()]
            
            if not keyword_list:
                st.error("유효한 키워드를 입력해주세요.")
            else:
                # Chrome 옵션 설정
                chrome_options = Options()
                chrome_options.add_argument("--headless")  # 헤드리스 모드

                # WebDriver 초기화
                driver = webdriver.Chrome(options=chrome_options)

                # 결과를 저장할 리스트 초기화
                results_list = []
                keyword_types = {}  # 키워드 유형을 저장할 딕셔너리
                smartblock_keywords = {}  # 스마트블럭 키워드와 연관 키워드를 저장할 딕셔너리
                
                # 실시간 결과 표시를 위한 placeholder
                result_placeholder = st.empty()

                # 진행 상황 표시를 위한 progress bar
                progress_bar = st.progress(0)

                # 각 키워드에 대해 검색 수행
                for i, keyword in enumerate(keyword_list):
                    # 검색 페이지로 이동
                    driver.get(f"https://search.naver.com/search.naver?ssc=tab.nx.all&where=nexearch&sm=tab_jum&query={keyword}")

                    try:
                        keyword_type = ''
                        is_knowledge_snippet = False
                        is_smartblock = False

                        # 지식스니펫 확인
                        try:
                            knowledge_snippet = driver.find_element(By.CSS_SELECTOR, '.source_box .txt.elss')
                            is_knowledge_snippet = True
                        except:
                            pass

                        # 스마트블럭 확인
                        try:
                            smartblock_research = driver.find_element(By.CSS_SELECTOR, '.BZppu7wV32H2scXPRUVx.fds-info-inner-text')
                            is_smartblock = True
                        except:
                            pass

                        # 키워드 유형 결정
                        if is_knowledge_snippet and is_smartblock:
                            keyword_type = 'both'
                        elif is_knowledge_snippet:
                            keyword_type = 'knowledge_snippet'
                        elif is_smartblock:
                            keyword_type = 'smartblock'

                        # 스마트블럭 연관 키워드 추출 (스마트블럭이거나 둘 다인 경우)
                        if is_smartblock:
                            related_keywords = driver.find_elements(By.CSS_SELECTOR, '.THdc2aWT6Ffq9q29WTab.fds-comps-keyword-chip-text.PJMOS12GUdMwfrs67QQn')
                            smartblock_keywords[keyword] = [k.text for k in related_keywords]

                        # 키워드 유형 저장
                        keyword_types[keyword] = keyword_type

                        # 블로그 탭 클릭
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '.flick_bx:nth-of-type(3) > a'))
                        ).click()
                        
                        # 무한스크롤 처리
                        last_height = driver.execute_script("return document.body.scrollHeight")
                        while True:
                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                            time.sleep(random.uniform(1, 1.5))
                            new_height = driver.execute_script("return document.body.scrollHeight")
                            if new_height == last_height:
                                break
                            last_height = new_height

                        # 블로그 순위 체크
                        blog_ids = driver.find_elements(By.CSS_SELECTOR, '.user_info a')
                        results = {j: '' for j in range(1, 16)}  # 모든 순위를 빈 문자열로 초기화
                        for blog_id in blog_ids:
                            href = blog_id.get_attribute('href')
                            extracted_id = href.split('/')[-1]
                            if extracted_id in dongju_id_list:
                                index = dongju_id_list.index(extracted_id) + 1
                                if index <= 15:  # 15위까지만 저장
                                    results[index] = extracted_id

                        # 검색량 조회
                        pc_volume, mobile_volume = get_search_volume(keyword)

                        # 결과 리스트에 추가
                        row = {'키워드': keyword, 'M': mobile_volume, 'P': pc_volume}
                        row.update(results)
                        results_list.append(row)

                        # 실시간으로 결과 표시
                        df = pd.DataFrame(results_list)
                        
                        # 키워드 열에만 배경색 적용
                        styled_df = df.style.apply(lambda row: [color_keyword(val, keyword_types, row['키워드']) if idx == 0 else '' for idx, val in enumerate(row)], axis=1)
                        
                        result_placeholder.dataframe(styled_df, width=1000)  # 너비 조정

                        # 진행 상황 업데이트
                        progress_bar.progress((i + 1) / len(keyword_list))

                    except Exception as e:
                        st.error(f"키워드 '{keyword}' 검색 중 오류 발생: {str(e)}")

                driver.quit()

                # Progress bar 제거
                progress_bar.empty()

                # 스마트블럭 키워드 및 연관 키워드 표시
                if smartblock_keywords:
                    st.subheader("스마트블럭 키워드 및 연관 키워드")
                    smartblock_data = []
                    for keyword, related_keywords in smartblock_keywords.items():
                        smartblock_data.append({
                            "키워드": keyword,
                            "연관 키워드": ", ".join(related_keywords)
                        })
                    smartblock_df = pd.DataFrame(smartblock_data)
                    st.dataframe(smartblock_df, width=1000)

                # 엑셀 다운로드 버튼
                excel_data = create_excel(df, keyword_types, smartblock_keywords)
                st.download_button(
                    label="엑셀 다운로드",
                    data=excel_data,
                    file_name="search_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    st.info("'순위 확인' 버튼을 클릭해서 검색 결과를 확인하세요.")

elif selected_tab == "구글":
    st.title("구글 순위 체크")
    st.write("준비중...")
